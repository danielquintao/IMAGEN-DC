from openpyxl import load_workbook,Workbook
import re
import numpy as np

def check_table(f, known_words_set):
        """read table and return columns containing undefined values
        (this is simpler than doing a similar check for each cell)
        """
        wb = load_workbook(f)
        s = wb.active
        irregular_cols = []
        for col in s.iter_cols(values_only=True):
                for i,val in enumerate(col):
                        if i == 0: # header / question name
                                continue
                        # value is ok if: it is a number, None or trivial, a stringifyied number, or an elem of known_words 
                        elif isinstance(val, (int, float)):
                                continue
                        elif val is None:
                                continue
                        elif isinstance(val, str):
                                val = val.strip()
                                if val in known_words_set:
                                        continue
                                try:
                                        val = float(val)
                                except ValueError:
                                        # tag column as irregular and break loop (go to next column)
                                        irregular_cols.append(col[0])
                                        break               
        return irregular_cols                         

def get_table(f):
        """takes filename and return list of list with the contents of all the rows
        (headers/question names are also processed)
        """
        wb = load_workbook(f)
        s = wb.active
        def iter_rows(s):
                empty_space = re.compile('\\s+')
                for row in s.iter_rows():
                        if not(row[0].value is None):
                                if isinstance(row[0].value, str) and empty_space.fullmatch(row[0].value):
                                        continue
                                yield [cell.value if not(cell.value is None) else np.inf for cell in row]
        #out = [[c.value for c in l] for l in s.rows]
        out = list(iter_rows(s))
        #Post process 0 : Change [0][0] to ID
        out[0][0] = "ID"
        #Post process 1 : remove doubles (duplicated lines -> happens in some files)
        # def fusion(l1,l2):
        #         l = l1
        #         for i in range(len(l)):
        #                 if l[i] is None:
        #                         l[i] = l2[i]
        #         return l
        # to_del = []
        for i in range(len(out)-1):
                assert out[i][0] != out[i+1][0]
                # if out[i][0] == out[i+1][0]:
                        # out[i] = fusion(out[i],out[i+1])
                        # to_del += [i+1]
        # to_del.sort(reverse=True)
        # for i in to_del:
        #         del out[i]
        #Post process 2 : add timestep and questionnaire name to labels and cut everything after the first space
        l = re.split("/",f)
        l = re.split("\.",l[-1])  
        if re.search('_(bas|fu[1-3])', l[0][-4:]):
                l[0][-4] = '-' # fix common error of naming files with _xxx instead of -xxx              
        l2 = re.split('-',l[0])
        # create a time prefix to be used later by find_matching_columns:
        # (1-3 for baseline, FU1, FU2, FU3; and 'C' as in 'Constant' for files with no time suffix)
        time = ['bas','fu1','fu2','fu3'].index(l2[1]) if len(l2) == 2 else 'C'
        empty_space = re.compile('\\s+')
        for i in range(1, len(out[0])): #starting from 1 -- don't change ID
                name = str(out[0][i])
                name = name.strip()
                name = empty_space.subn('_', name)[0]
                name = re.subn('³','3',name)[0]
                out[0][i] = str(time)+"_"+name+"_"+l2[0]
        if time == 'C':
                print("loaded table {} : ({},{}) (constant on time)".format(f,len(out),len(out[0])))
        else:
                print("loaded table {} : ({},{}) timestep {}".format(f,len(out),len(out[0]),time))
        return out
                
def find_matching_columns(t, conv):
        """Find columns corresponding to the same question on different time steps
        and align on the individuals (also process categorical values)
        conv is a dict of prefixed terms (str) as key and the associated numeric
        placeholder (value)"""
        def match_form(n): # n (str) is the question name in the covention [0-3|'C']_questionname_filename
                           # (no problem if there are underscores inside questionname or filename)
                time_prefix = int(n[0]) if n[0] != 'C' else 'C'
                return n[2:].lower(),time_prefix
        #Get matching names
        names = t[0]
        d_names = {} #Match dictionary
        # match_names = [] # Names matched with others -- debug
        for index,n in enumerate(names):
                if index != 0:
                        m_form,time = match_form(n)
                else: # index == 0 -> n is an ID
                        m_form,time = n,-1  # -1 -> IDs will be inserted as 'fu3' in the final tensor
                try:
                        if isinstance(time, int):
                                d_names[m_form].append((index,time))
                        else: # time == 'C', add constant values to all time points
                                assert time == 'C'
                                d_names[m_form].append((index,0))
                                d_names[m_form].append((index,1))
                                d_names[m_form].append((index,2))
                                d_names[m_form].append((index,3))
                        # match_names.append(n) -- debug
                except KeyError:
                        if isinstance(time, int):
                                d_names[m_form] = [(index,time)]
                        else: # time == 'C', add constant values to all time points
                                assert time == 'C'
                                d_names[m_form] = [(index,0)]
                                d_names[m_form].append((index,1))
                                d_names[m_form].append((index,2))
                                d_names[m_form].append((index,3))
        #print(d_names)
        #print(match_names,len(match_names))
        #Add dimensions
        nb_timesteps = 4
        nb_ind = len(t) # includes the header 'ID' despite the name
        nb_questions = len(d_names) # includes the 'ID' though not a question
        questions = sorted(list(d_names))
        tf = np.full((nb_timesteps,nb_ind,nb_questions),np.inf)
        for index_ind in range(1,nb_ind):
                for index_q in range(nb_questions):
                        q_name = questions[index_q]
                        indexs_tab = d_names[q_name]
                        for index,time in indexs_tab:
                                value = t[index_ind][index]
                                if value is None:
                                        value = np.inf
                                try:
                                        tf[time,index_ind,index_q] = value # NOTE it works even if value is a 'stringifyied' float like '0.123'
                                except ValueError:
                                        #Try to convert it
                                        try:
                                                value = int(value)
                                                tf[time,index_ind,index_q] = value
                                        except ValueError:
                                                try:
                                                        assert isinstance(value, str)
                                                        value = conv[value.strip()] # strip --> remove blank spaces at extremities ('B+ ' -> 'B+')
                                                        tf[time,index_ind,index_q] = value
                                                except KeyError:
                                                        # print("error : '{}' not understood".format(value))
                                                        value = np.nan #Error value
                                                        tf[time,index_ind,index_q] = value
        #tf = np.delete(tf,0,axis=2)
        tf = np.delete(tf,0,axis=1)
        timesteps = np.array(["bas","fu1","fu2","fu3"]).astype(str)
        ind = np.array(t)[:,0]
        ind = np.delete(ind,0)
        questions = np.array(questions).astype(str)
        return tf,timesteps,ind,questions
